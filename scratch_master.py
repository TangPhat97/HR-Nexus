import pandas as pd
from openpyxl import load_workbook
import json
path = 'o:/clean-architecture-canonical/LD REPORT V2/LD REPORT V2/file excel tren gg sheet/L&D - 2026 MASTER DATA - TEST v4.xlsx'
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb['Danh mục nhân viên']
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
df = pd.DataFrame(rows[1:], columns=headers)
cols = [df.columns[0], df.columns[1], df.columns[2]]
print('Search for thao.lam@gonsa.com.vn:')
print(json.dumps(df[df.eq('thao.lam@gonsa.com.vn').any(axis=1)][cols].to_dict('records')))
print('Search for chuong.le@gonsa.com.vn:')
print(json.dumps(df[df.eq('chuong.le@gonsa.com.vn').any(axis=1)][cols].to_dict('records')))
print('Search for 260311-03361:')
print(json.dumps(df[df.eq('260311-03361').any(axis=1)][cols].to_dict('records')))
