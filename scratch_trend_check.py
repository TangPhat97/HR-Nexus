"""Scratch: check raw data for Hoi nhap NV moi 2026 attendance."""
import openpyxl
from pathlib import Path

XLSX = Path(r"O:\clean-architecture-canonical\LD REPORT V2\LD REPORT V2\file excel tren gg sheet\LD REPORT TEST 2026 v8.xlsx")

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

print("=== SHEETS ===")
for s in wb.sheetnames:
    print(f"  {s}")

# Find raw data sheet
for sn in wb.sheetnames:
    lower = sn.lower()
    if "hoc vien" in lower or "raw" in lower or "data raw" in lower or "học viên" in lower:
        ws = wb[sn]
        rows_list = list(ws.iter_rows(min_row=1, values_only=True))
        headers = rows_list[0] if rows_list else []
        print(f"\n=== Sheet: {sn} ===")
        print(f"  Headers: {list(headers[:25])}")
        print(f"  Total data rows: {len(rows_list) - 1}")

        # Find attendance column index
        att_col = None
        course_col = None
        for i, h in enumerate(headers):
            hs = str(h).lower() if h else ""
            if "điểm danh" in hs or "diem danh" in hs or "trạng thái" in hs or "attendance" in hs:
                att_col = i
            if "khóa" in hs or "khoa" in hs or "tên lớp" in hs or "ten lop" in hs or "course" in hs:
                if course_col is None:
                    course_col = i

        print(f"  Attendance col index: {att_col}")
        print(f"  Course col index: {course_col}")

        # Count Hoi nhap records
        hoi_nhap_rows = []
        for row_idx, r in enumerate(rows_list[1:], start=2):
            row_str = " ".join(str(v) for v in r if v)
            if "HOINHAP" in row_str.upper() or "Hội nhập" in row_str or "hoi nhap" in row_str.lower():
                att_val = str(r[att_col]).strip() if att_col is not None and att_col < len(r) else "?"
                course_val = str(r[course_col]).strip() if course_col is not None and course_col < len(r) else "?"
                hoi_nhap_rows.append((row_idx, course_val[:40], att_val))

        print(f"\n  Hoi nhap total: {len(hoi_nhap_rows)}")
        co_mat = sum(1 for _, _, att in hoi_nhap_rows if att == "Có mặt")
        vang = sum(1 for _, _, att in hoi_nhap_rows if "Vắng" in att or "vang" in att.lower())
        print(f"  Co mat: {co_mat}")
        print(f"  Vang: {vang}")
        print(f"  Other: {len(hoi_nhap_rows) - co_mat - vang}")

        print("\n  Detail:")
        for row_num, course, att in hoi_nhap_rows:
            print(f"    Row {row_num}: [{att}] {course}")

# Also check the trend output sheet
for sn in wb.sheetnames:
    if "xu hướng" in sn.lower() or "xu huong" in sn.lower() or "trend" in sn.lower():
        ws = wb[sn]
        print(f"\n=== OUTPUT Sheet: {sn} ===")
        for r in ws.iter_rows(min_row=1, max_row=30, values_only=True):
            vals = [v for v in r if v is not None]
            if vals:
                print(f"  {vals[:17]}")

wb.close()
