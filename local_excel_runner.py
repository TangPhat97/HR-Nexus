from __future__ import annotations

import argparse
import json
import shutil
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from local_fact_builder import build_local_training_sync
from transform import AnalyticsInputs, SHEET_NAMES, transform_data

Logger = Callable[[str], None]

EMPLOYEES_SHEET = "Danh mục nhân viên"
COURSES_SHEET = "Danh mục khóa học"
TRAINING_SESSIONS_SHEET = "Lớp đào tạo"
RAW_PARTICIPANTS_SHEET = "Data raw học viên"
TRAINING_RECORDS_SHEET = "Dữ liệu đào tạo"
QUEUE_JOBS_SHEET = "Hàng đợi xử lý"
QA_RESULTS_SHEET = "Kết quả QA"
STAGING_HR_SHEET = "Staging nhân sự"
STAGING_COURSES_SHEET = "Staging khóa học"
STAGING_TRAINING_SHEET = "Staging đào tạo"
CONFIG_SHEET = "Cấu hình hệ thống"

SOURCE_SHEETS = (
    EMPLOYEES_SHEET,
    COURSES_SHEET,
    TRAINING_SESSIONS_SHEET,
    RAW_PARTICIPANTS_SHEET,
    TRAINING_RECORDS_SHEET,
    QUEUE_JOBS_SHEET,
    QA_RESULTS_SHEET,
    STAGING_HR_SHEET,
    STAGING_COURSES_SHEET,
    STAGING_TRAINING_SHEET,
    CONFIG_SHEET,
)

OUTPUT_SHEETS = tuple(SHEET_NAMES.values())


DEFAULT_WORKBOOK_PATH = (
    Path(__file__).resolve().parent
    / "file excel tren gg sheet"
    / "L&D - 2026 MASTER DATA - TEST v4.xlsx"
)

HEADER_ALIASES = {
    EMPLOYEES_SHEET: {
        "ma_nhan_vien": "emp_id",
        "ho_ten": "full_name",
        "email": "email",
        "tinh_trang_nhan_su": "employment_status",
        "cong_ty": "company",
        "phong_ban": "department",
        "khoi": "division",
        "chuc_danh": "job_title",
        "ngach": "grade",
        "cap_bac": "level",
        "trang_thai_dong": "row_status",
        "dia_ban": "region",
    },
    COURSES_SHEET: {
        "ma_khoa_hoc": "course_id",
        "ten_khoa_hoc": "course_name",
        "nhom_khoa_hoc": "course_category",
        "nen_tang": "platform",
        "thoi_luong_gio": "duration_hours",
        "chi_phi_nguoi": "cost_per_pax",
        "mo_ta_khoa_hoc": "course_description",
        "trang_thai_dong": "row_status",
        "loai_hinh_dao_tao": "delivery_type",
        "hinh_thuc_mac_dinh": "training_format_default",
        "don_vi_dao_tao": "training_unit",
        "doi_tuong_dao_tao": "target_audience",
        "pham_vi_cong_ty": "company_scope",
    },
    TRAINING_SESSIONS_SHEET: {
        "ma_lop_dao_tao": "session_id",
        "ma_lop": "session_code",
        "ma_khoa_hoc": "course_id",
        "ten_khoa_hoc": "course_name",
        "loai_chuong_trinh": "program_type",
        "trong_ngoai_ke_hoach": "plan_scope",
        "loai_hinh_dao_tao": "delivery_type",
        "hinh_thuc_dao_tao": "training_format",
        "dia_diem_dao_tao": "location",
        "don_vi_dao_tao": "training_unit",
        "doi_tuong_dao_tao": "target_audience",
        "pham_vi_cong_ty": "company_scope",
        "ngay_dao_tao": "training_date",
        "thang_dao_tao": "training_month",
        "so_luong_lop": "class_count",
        "so_luong_hv_dang_ky": "registered_count",
        "so_luong_hv_thuc_te": "actual_count",
        "thoi_luong_gio": "duration_hours",
        "tong_gio_dao_tao": "total_hours",
        "tong_gio_dao_tao_man_hour": "total_man_hours",
        "chi_phi_dao_tao_du_kien": "estimated_cost",
        "hoc_phi_chi_phi_giang_vien": "instructor_cost",
        "chi_phi_to_chuc": "organization_cost",
        "chi_phi_binh_quan_nguoi": "avg_cost_per_pax",
        "ghi_nhan_chi_phi": "expense_status",
        "phieu_yeu_cau_dao_tao_email": "iso_request_reference",
        "ke_hoach_du_tru_chi_phi": "iso_budget_reference",
        "danh_sach_diem_danh": "iso_attendance_evidence",
        "chung_tu_khac": "iso_other_reference",
        "trang_thai_dong": "row_status",
    },
    RAW_PARTICIPANTS_SHEET: {
        "ma_raw": "raw_id",
        "ma_lop_dao_tao": "session_id",
        "ma_lop": "session_code",
        "ma_khoa_hoc": "course_id",
        "ten_khoa_hoc": "course_name",
        "ngay_dao_tao": "training_date",
        "ma_nhan_vien": "emp_id",
        "ho_ten": "full_name",
        "email": "email",
        "trang_thai_diem_danh": "attendance_status",
        "diem": "score",
        "hai_long": "satisfaction",
        "muc_do_lien_quan": "relevance",
        "nps": "nps",
        "ap_dung_vao_cong_viec": "applied_on_job",
        "nhan_xet_quan_ly": "manager_comment",
        "ma_bam_dong": "source_row_hash",
        "cap_nhat_luc": "updated_at",
        "cap_nhat_boi": "updated_by",
        "trang_thai_dong": "row_status",
        "ghi_chu": "notes",
    },
    TRAINING_RECORDS_SHEET: {
        "ma_ban_ghi": "record_id",
        "loai_nguon": "source_type",
        "ma_batch": "batch_id",
        "ma_nhan_vien": "emp_id",
        "ho_ten": "full_name",
        "email": "email",
        "tinh_trang_nhan_su": "employment_status",
        "cong_ty": "company",
        "khoi": "division",
        "phong_ban": "department",
        "chuc_danh": "job_title",
        "ngach": "grade",
        "cap_bac": "level",
        "ma_khoa_hoc": "course_id",
        "ten_khoa_hoc": "course_name",
        "nhom_khoa_hoc": "course_category",
        "nen_tang": "platform",
        "thoi_luong_gio": "duration_hours",
        "chi_phi_nguoi": "cost_per_pax",
        "ngay_dao_tao": "training_date",
        "hinh_thuc_dao_tao": "training_format",
        "trang_thai_diem_danh": "attendance_status",
        "diem": "score",
        "hai_long": "satisfaction",
        "muc_do_lien_quan": "relevance",
        "nps": "nps",
        "ap_dung_vao_cong_viec": "applied_on_job",
        "nhan_xet_quan_ly": "manager_comment",
        "tao_luc": "created_at",
        "tao_boi": "created_by",
        "cap_nhat_luc": "updated_at",
        "cap_nhat_boi": "updated_by",
        "nam_archive": "archive_year",
        "trang_thai_dong": "row_status",
        "ma_bam_dong": "source_row_hash",
        "trang_thai_qa": "qa_status",
        "metadata": "metadata_json",
        "ma_lop_dao_tao": "session_id",
        "ma_lop": "session_code",
        "thang_dao_tao": "training_month",
        "dia_diem_dao_tao": "location",
        "loai_hinh_dao_tao": "delivery_type",
        "don_vi_dao_tao": "training_unit",
        "doi_tuong_dao_tao": "target_audience",
        "pham_vi_cong_ty": "company_scope",
        "so_luong_lop": "class_count",
        "so_luong_hv_dang_ky": "registered_count",
        "so_luong_hv_thuc_te": "actual_count",
        "chi_phi_dao_tao_du_kien": "estimated_cost",
        "hoc_phi_chi_phi_giang_vien": "instructor_cost",
        "chi_phi_to_chuc": "organization_cost",
        "phieu_yeu_cau_dao_tao_email": "iso_request_reference",
        "ke_hoach_du_tru_chi_phi": "iso_budget_reference",
        "danh_sach_diem_danh": "iso_attendance_evidence",
        "chung_tu_khac": "iso_other_reference",
        "dia_ban": "region",
    },
    QUEUE_JOBS_SHEET: {
        "loai_cong_viec": "job_type",
        "trang_thai": "status",
        "du_lieu_payload": "payload_json",
        "checkpoint": "checkpoint_json",
        "cap_nhat_luc": "updated_at",
        "tao_luc": "created_at",
    },
    QA_RESULTS_SHEET: {
        "muc_do": "severity",
        "trang_thai": "status",
    },
    STAGING_HR_SHEET: {
        "trang_thai_dong": "row_status",
    },
    STAGING_COURSES_SHEET: {
        "trang_thai_dong": "row_status",
    },
    STAGING_TRAINING_SHEET: {
        "trang_thai_dong": "row_status",
    },
    CONFIG_SHEET: {
        "ma_cau_hinh": "config_key",
        "gia_tri": "config_value",
    },
}

REQUIRED_COLUMNS = {
    EMPLOYEES_SHEET: ("emp_id", "department", "row_status"),
    COURSES_SHEET: ("course_name", "row_status"),
    TRAINING_SESSIONS_SHEET: ("session_id", "course_name", "training_date", "row_status"),
    RAW_PARTICIPANTS_SHEET: ("session_id", "emp_id", "attendance_status", "row_status"),
    TRAINING_RECORDS_SHEET: (
        "emp_id",
        "course_name",
        "training_date",
        "attendance_status",
        "row_status",
    ),
    QUEUE_JOBS_SHEET: ("job_type", "status"),
    QA_RESULTS_SHEET: ("severity", "status"),
    STAGING_HR_SHEET: ("row_status",),
    STAGING_COURSES_SHEET: ("row_status",),
    STAGING_TRAINING_SHEET: ("row_status",),
    CONFIG_SHEET: ("config_key", "config_value"),
}


@dataclass(frozen=True)
class SheetStatus:
    expected_name: str
    actual_name: str
    row_count: int
    recognized_columns: tuple[str, ...]
    missing_columns: tuple[str, ...]
    is_missing: bool = False


@dataclass(frozen=True)
class WorkbookInspection:
    workbook_path: Path
    fiscal_year: str
    source_statuses: tuple[SheetStatus, ...]
    warnings: tuple[str, ...]

    @property
    def source_row_counts(self) -> dict[str, int]:
        return {status.expected_name: status.row_count for status in self.source_statuses}


@dataclass(frozen=True)
class LocalRefreshResult:
    workbook_path: Path
    backup_path: Path | None
    fiscal_year: str
    inspection: WorkbookInspection
    output_shapes: dict[str, tuple[int, int]]


def fetch_data(workbook_path: str | Path, logger: Logger | None = None) -> AnalyticsInputs:
    path = _resolve_workbook_path(workbook_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        tables, statuses, warnings = _load_tables(workbook)
        _emit(logger, f"Da doc workbook local: {path}")
        for status in statuses:
            _emit(logger, f"- {status.expected_name}: {status.row_count} dong")
        for warning in warnings:
            _emit(logger, f"Canh bao: {warning}")
        config_map = _build_config_map(tables[CONFIG_SHEET])
        raw_sync_summary = _build_raw_sync_summary(tables[QUEUE_JOBS_SHEET])
        return AnalyticsInputs(
            training_records=tables[TRAINING_RECORDS_SHEET],
            employees=tables[EMPLOYEES_SHEET],
            training_sessions=tables[TRAINING_SESSIONS_SHEET],
            raw_participants=tables[RAW_PARTICIPANTS_SHEET],
            queue_jobs=tables[QUEUE_JOBS_SHEET],
            qa_results=tables[QA_RESULTS_SHEET],
            staging_hr=tables[STAGING_HR_SHEET],
            staging_courses=tables[STAGING_COURSES_SHEET],
            staging_training=tables[STAGING_TRAINING_SHEET],
            config_map=config_map,
            raw_sync_summary=raw_sync_summary,
        )
    finally:
        workbook.close()


def inspect_workbook(
    workbook_path: str | Path,
    fiscal_year: str | None = None,
    logger: Logger | None = None,
) -> WorkbookInspection:
    path = _resolve_workbook_path(workbook_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        tables, statuses, warnings = _load_tables(workbook)
        config_map = _build_config_map(tables[CONFIG_SHEET])
        resolved_year = str(fiscal_year or config_map.get("CURRENT_FISCAL_YEAR") or datetime.now().year)
        target_lookup = _sheet_lookup(workbook)
        target_warnings = [
            f"Thieu sheet output '{sheet_name}', local runner se tao moi khi ghi."
            for sheet_name in OUTPUT_SHEETS
            if _normalize_token(sheet_name) not in target_lookup
        ]
        combined_warnings = tuple(list(warnings) + target_warnings)
        inspection = WorkbookInspection(
            workbook_path=path,
            fiscal_year=resolved_year,
            source_statuses=tuple(statuses),
            warnings=combined_warnings,
        )
        _emit(logger, f"Workbook: {path}")
        _emit(logger, f"Nam bao cao: {inspection.fiscal_year}")
        for status in inspection.source_statuses:
            line = f"- {status.expected_name}: {status.row_count} dong"
            if status.missing_columns:
                line += f" | thieu cot: {', '.join(status.missing_columns)}"
            _emit(logger, line)
        for warning in inspection.warnings:
            _emit(logger, f"Canh bao: {warning}")
        return inspection
    finally:
        workbook.close()


def write_data(
    workbook_path: str | Path,
    outputs: dict[str, list[list[Any]]],
    source_updates: dict[str, pd.DataFrame] | None = None,
    backup: bool = True,
    logger: Logger | None = None,
) -> Path | None:
    path = _resolve_workbook_path(workbook_path)
    backup_path = _create_backup(path) if backup else None
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        lookup = _sheet_lookup(workbook)
        for sheet_name, frame in (source_updates or {}).items():
            worksheet = _ensure_target_sheet(workbook, lookup, sheet_name)
            _emit(logger, f"Ghi nguon: {sheet_name} ({len(frame)} dong)")
            _write_frame_body(worksheet, frame, sheet_name)
        for sheet_name, matrix in outputs.items():
            worksheet = _ensure_target_sheet(workbook, lookup, sheet_name)
            _emit(logger, f"Ghi sheet: {sheet_name} ({len(matrix)} dong)")
            _write_matrix(worksheet, matrix)
            _apply_post_write_formatting(sheet_name, worksheet)
        if hasattr(workbook, "calculation") and workbook.calculation is not None:
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
        workbook.save(path)
        if backup_path is not None:
            _emit(logger, f"Da tao backup: {backup_path}")
        _emit(logger, f"Da cap nhat workbook: {path}")
        return backup_path
    finally:
        workbook.close()


def run_local_refresh(
    workbook_path: str | Path,
    fiscal_year: str | None = None,
    backup: bool = True,
    logger: Logger | None = None,
) -> LocalRefreshResult:
    inspection = inspect_workbook(workbook_path, fiscal_year=fiscal_year, logger=logger)
    path = _resolve_workbook_path(workbook_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        tables, _, _ = _load_tables(workbook)
    finally:
        workbook.close()
    local_sync = build_local_training_sync(
        employees=tables[EMPLOYEES_SHEET],
        courses=tables[COURSES_SHEET],
        training_sessions=tables[TRAINING_SESSIONS_SHEET],
        raw_participants=tables[RAW_PARTICIPANTS_SHEET],
        existing_records=tables[TRAINING_RECORDS_SHEET],
        logger=logger,
    )
    inputs = AnalyticsInputs(
        training_records=local_sync.training_records,
        employees=tables[EMPLOYEES_SHEET],
        training_sessions=tables[TRAINING_SESSIONS_SHEET],
        raw_participants=local_sync.raw_participants,
        queue_jobs=tables[QUEUE_JOBS_SHEET],
        qa_results=tables[QA_RESULTS_SHEET],
        staging_hr=tables[STAGING_HR_SHEET],
        staging_courses=tables[STAGING_COURSES_SHEET],
        staging_training=tables[STAGING_TRAINING_SHEET],
        config_map=_build_config_map(tables[CONFIG_SHEET]),
        raw_sync_summary={"failed_rows": local_sync.failed_rows},
    )
    _emit(logger, "Dang tinh toan 8 sheet bao cao bang pandas...")
    outputs = transform_data(inputs, fiscal_year=inspection.fiscal_year)
    backup_path = write_data(
        workbook_path,
        outputs,
        source_updates={
            RAW_PARTICIPANTS_SHEET: local_sync.raw_participants,
            TRAINING_RECORDS_SHEET: local_sync.training_records,
            EMPLOYEES_SHEET: local_sync.employees,
        },
        backup=backup,
        logger=logger,
    )
    output_shapes = {
        sheet_name: (len(matrix), max((len(row) for row in matrix), default=0))
        for sheet_name, matrix in outputs.items()
    }
    return LocalRefreshResult(
        workbook_path=_resolve_workbook_path(workbook_path),
        backup_path=backup_path,
        fiscal_year=inspection.fiscal_year,
        inspection=inspection,
        output_shapes=output_shapes,
    )


def _load_tables(workbook: Workbook) -> tuple[dict[str, pd.DataFrame], list[SheetStatus], list[str]]:
    lookup = _sheet_lookup(workbook)
    tables: dict[str, pd.DataFrame] = {}
    statuses: list[SheetStatus] = []
    warnings: list[str] = []
    for sheet_name in SOURCE_SHEETS:
        frame, status = _load_sheet(workbook, lookup, sheet_name)
        tables[sheet_name] = frame
        statuses.append(status)
        if status.is_missing:
            warnings.append(f"Khong tim thay sheet nguon '{sheet_name}'.")
        elif status.missing_columns:
            warnings.append(
                f"Sheet '{status.actual_name}' thieu cot can thiet: {', '.join(status.missing_columns)}."
            )
    # ── Dedup: phat hien + loai bo emp_id trung lap (Tang 1+2) ──
    tables[EMPLOYEES_SHEET], emp_dedup_warnings = _dedup_employees(tables[EMPLOYEES_SHEET])
    warnings.extend(emp_dedup_warnings)
    raw_rows = len(tables[RAW_PARTICIPANTS_SHEET])
    training_rows = len(tables[TRAINING_RECORDS_SHEET])
    config_rows = len(tables[CONFIG_SHEET])
    if training_rows == 0:
        warnings.append(
            "Sheet 'Dữ liệu đào tạo' dang rong. POC local hien chi tinh analytics tu fact table nay."
        )
    if raw_rows > 0 and training_rows == 0:
        warnings.append(
            "Da co 'Data raw học viên' nhung chua co 'Dữ liệu đào tạo'; local runner chua thay the raw-sync GAS."
        )
    if config_rows > 1000:
        warnings.append(
            "Sheet 'Cấu hình hệ thống' co so dong bat thuong lon. Local runner van doc duoc, "
            "nhung nen don dep du lieu rac tu loi auto-append cu."
        )
    return tables, statuses, warnings


def _dedup_employees(frame: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Detect and remove duplicate emp_id entries, keeping the last row.

    Tang 1: Generate warnings listing every duplicate with row numbers.
    Tang 2: Auto-deduplicate by keeping the last occurrence (highest __rowNumber).
    """
    warnings: list[str] = []
    if frame.empty or "emp_id" not in frame.columns:
        return frame, warnings

    # Normalize emp_id for comparison
    emp_ids = frame["emp_id"].astype("string").str.strip()
    has_id_mask = emp_ids.ne("") & emp_ids.notna() & emp_ids.ne("<NA>")

    if not has_id_mask.any():
        return frame, warnings

    working = frame.loc[has_id_mask].copy()
    working["_emp_id_clean"] = emp_ids[has_id_mask]

    # Find emp_ids that appear more than once
    counts = working["_emp_id_clean"].value_counts()
    dup_ids = counts[counts > 1]

    if dup_ids.empty:
        return frame, warnings

    # ── Tang 1: Canh bao chi tiet ──
    for emp_id in dup_ids.index:
        group = working.loc[working["_emp_id_clean"] == emp_id]
        name = ""
        if "full_name" in group.columns:
            names = group["full_name"].dropna()
            name = str(names.iloc[0]) if not names.empty else ""
        rows = group["__rowNumber"].tolist()
        row_str = ", ".join(str(int(r)) for r in rows if pd.notna(r))
        kept = max(r for r in rows if pd.notna(r))
        warnings.append(
            f"Ma NV '{emp_id}' ({name}) trung lap tai dong {row_str}. "
            f"He thong giu dong {int(kept)} (moi nhat), bo dong con lai."
        )

    # ── Tang 2: Tu dong dedup — giu dong cuoi cung theo __rowNumber ──
    no_id_rows = frame.loc[~has_id_mask]
    has_id_rows = frame.loc[has_id_mask].copy()
    has_id_rows["_emp_id_clean"] = emp_ids[has_id_mask]
    deduped = has_id_rows.drop_duplicates(subset=["_emp_id_clean"], keep="last")
    deduped = deduped.drop(columns=["_emp_id_clean"])

    result = pd.concat([deduped, no_id_rows], ignore_index=False).sort_index()
    total_removed = len(frame) - len(result)
    if total_removed > 0:
        warnings.append(
            f"Da loai {total_removed} dong trung lap tu 'Danh muc nhan vien'."
        )
    return result, warnings


def _load_sheet(
    workbook: Workbook,
    lookup: dict[str, str],
    expected_name: str,
) -> tuple[pd.DataFrame, SheetStatus]:
    normalized_name = _normalize_token(expected_name)
    if normalized_name not in lookup:
        empty = pd.DataFrame({column: pd.Series(dtype="object") for column in REQUIRED_COLUMNS[expected_name]})
        status = SheetStatus(
            expected_name=expected_name,
            actual_name=expected_name,
            row_count=0,
            recognized_columns=tuple(),
            missing_columns=tuple(REQUIRED_COLUMNS[expected_name]),
            is_missing=True,
        )
        return empty, status
    actual_name = lookup[normalized_name]
    worksheet = workbook[actual_name]
    frame = _worksheet_to_frame(worksheet)
    alias_map = HEADER_ALIASES[expected_name]
    rename_map = {
        column: alias_map[_normalize_token(column)]
        for column in frame.columns
        if _normalize_token(column) in alias_map
    }
    frame = frame.rename(columns=rename_map)
    missing_columns = tuple(column for column in REQUIRED_COLUMNS[expected_name] if column not in frame.columns)
    status = SheetStatus(
        expected_name=expected_name,
        actual_name=actual_name,
        row_count=len(frame),
        recognized_columns=tuple(sorted(set(rename_map.values()))),
        missing_columns=missing_columns,
        is_missing=False,
    )
    return frame, status


def _worksheet_to_frame(worksheet: Worksheet) -> pd.DataFrame:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    header_row = [_sanitize_header_cell(cell) for cell in rows[0]]
    if not any(header_row):
        return pd.DataFrame()
    body_rows: list[list[Any]] = []
    row_numbers: list[int] = []
    width = len(header_row)
    for row_number, row in enumerate(rows[1:], start=2):
        values = list(row[:width])
        if len(values) < width:
            values.extend([None] * (width - len(values)))
        if all(_is_blank_cell(value) for value in values):
            continue
        body_rows.append(values)
        row_numbers.append(row_number)
    if not body_rows:
        frame = pd.DataFrame(columns=header_row)
    else:
        frame = pd.DataFrame(body_rows, columns=header_row)
    frame["__rowNumber"] = pd.Series(row_numbers, dtype="Int64")
    return frame


def _build_config_map(frame: pd.DataFrame) -> dict[str, Any]:
    if frame.empty or "config_key" not in frame.columns or "config_value" not in frame.columns:
        return {}
    config_frame = frame.copy()
    config_frame["config_key"] = config_frame["config_key"].astype("string").fillna("").str.strip()
    config_frame = config_frame.loc[config_frame["config_key"].ne("")]
    if config_frame.empty:
        return {}
    config_frame = config_frame.drop_duplicates(subset=["config_key"], keep="last")
    return {
        key: _normalize_config_value(value)
        for key, value in config_frame.set_index("config_key")["config_value"].to_dict().items()
    }


def _build_raw_sync_summary(frame: pd.DataFrame) -> dict[str, Any]:
    if frame.empty or "checkpoint_json" not in frame.columns:
        return {}
    working = frame.copy()
    working["job_type_key"] = working.get("job_type", pd.Series(dtype="string")).astype("string").fillna("").str.upper()
    working["updated_at_dt"] = pd.to_datetime(working.get("updated_at"), errors="coerce")
    working["checkpoint_dict"] = working["checkpoint_json"].map(_parse_json)
    sync_jobs = working.loc[working["job_type_key"].eq("SYNC_TRAINING_DATA")]
    candidates = sync_jobs if not sync_jobs.empty else working
    if candidates.empty:
        return {}
    latest = candidates.sort_values(["updated_at_dt", "__rowNumber"], na_position="last").tail(1)
    checkpoint = latest["checkpoint_dict"].iloc[0]
    if not isinstance(checkpoint, dict):
        return {}
    summary = checkpoint.get("raw_sync_summary")
    if isinstance(summary, dict):
        return summary
    failed_rows = checkpoint.get("failed_rows")
    if failed_rows is None:
        failed_rows = checkpoint.get("sync_runtime_failed_rows")
    return {"failed_rows": failed_rows or 0}


def _parse_json(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return dict(value)
    if value is None:
        return {}
    text = str(value).strip()
    if not text:
        return {}
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _normalize_config_value(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return value


def _write_matrix(worksheet: Worksheet, matrix: list[list[Any]]) -> None:
    row_count = len(matrix)
    col_count = max((len(row) for row in matrix), default=0)
    clear_rows = max(worksheet.max_row or 0, row_count)
    clear_cols = max(worksheet.max_column or 0, col_count)
    merged_ranges = [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
    for cell_range in merged_ranges:
        worksheet.unmerge_cells(cell_range)
    for row in worksheet.iter_rows(min_row=1, max_row=max(clear_rows, 1), min_col=1, max_col=max(clear_cols, 1)):
        for cell in row:
            cell.value = None
    for row_index, row in enumerate(matrix, start=1):
        for col_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=col_index, value=_excel_safe_value(value))
    for cell_range in merged_ranges:
        try:
            worksheet.merge_cells(cell_range)
        except ValueError:
            continue


def _write_frame_body(worksheet: Worksheet, frame: pd.DataFrame, sheet_name: str) -> None:
    header_row = [cell.value for cell in worksheet[1]]
    if not any(header_row):
        return
    alias_map = HEADER_ALIASES.get(sheet_name, {})
    data = {}
    for header in header_row:
        header_text = _sanitize_header_cell(header)
        normalized = _normalize_token(header_text)
        canonical = alias_map.get(normalized)
        if canonical and canonical in frame.columns:
            data[header_text] = frame[canonical]
        elif header_text in frame.columns:
            data[header_text] = frame[header_text]
        else:
            data[header_text] = pd.Series([""] * len(frame), dtype="object")
    body_frame = pd.DataFrame(data)
    clear_rows = max(worksheet.max_row or 1, len(body_frame) + 1)
    clear_cols = max(worksheet.max_column or len(header_row), len(header_row))
    for row in worksheet.iter_rows(min_row=2, max_row=max(clear_rows, 2), min_col=1, max_col=max(clear_cols, 1)):
        for cell in row:
            cell.value = None
    for row_index, row in enumerate(body_frame.itertuples(index=False, name=None), start=2):
        for col_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=col_index, value=_excel_safe_value(value))


def _excel_safe_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, pd.Timedelta):
        return str(value)
    if pd.isna(value):
        return None
    if hasattr(value, "item") and callable(value.item):
        try:
            return value.item()
        except Exception:
            return value
    return value


def _ensure_target_sheet(workbook: Workbook, lookup: dict[str, str], expected_name: str) -> Worksheet:
    normalized_name = _normalize_token(expected_name)
    actual_name = lookup.get(normalized_name)
    if actual_name is not None:
        return workbook[actual_name]
    worksheet = workbook.create_sheet(expected_name)
    lookup[normalized_name] = worksheet.title
    return worksheet


def _create_backup(workbook_path: Path) -> Path:
    backup_root = workbook_path.parent / "_local_backups"
    backup_root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = backup_root / f"{workbook_path.stem}__backup__{timestamp}{workbook_path.suffix}"
    shutil.copy2(workbook_path, backup_path)
    return backup_path


def _sheet_lookup(workbook: Workbook) -> dict[str, str]:
    return {_normalize_token(name): name for name in workbook.sheetnames}


def _resolve_workbook_path(workbook_path: str | Path) -> Path:
    path = Path(workbook_path).expanduser()
    if not path.is_absolute():
        path = (Path.cwd() / path).resolve()
    if not path.exists():
        raise FileNotFoundError(f"Khong tim thay file Excel: {path}")
    return path


def _normalize_token(value: Any) -> str:
    text = str(value or "").strip()
    text = text.replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii").lower()
    chars = [character if character.isalnum() else "_" for character in text]
    normalized = "".join(chars).strip("_")
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized


def _sanitize_header_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_blank_cell(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _emit(logger: Logger | None, message: str) -> None:
    if logger is not None:
        logger(message)


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run the HR-NEXUS local Excel analytics refresh.")
    parser.add_argument(
        "--workbook",
        default=str(DEFAULT_WORKBOOK_PATH),
        help="Path to the local Excel workbook.",
    )
    parser.add_argument(
        "--fiscal-year",
        default="",
        help="Optional fiscal year override. Defaults to config CURRENT_FISCAL_YEAR.",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Skip creating a timestamped backup before writing output sheets.",
    )
    parser.add_argument(
        "action",
        choices=("check", "run"),
        nargs="?",
        default="run",
        help="Use 'check' for a preflight inspection or 'run' to recalculate output sheets.",
    )
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        pass
    parser = _build_parser()
    args = parser.parse_args(list(argv) if argv is not None else None)
    workbook_path = args.workbook
    fiscal_year = args.fiscal_year or None
    if args.action == "check":
        inspect_workbook(workbook_path, fiscal_year=fiscal_year, logger=print)
        return 0
    result = run_local_refresh(
        workbook_path,
        fiscal_year=fiscal_year,
        backup=not args.no_backup,
        logger=print,
    )
    print(f"Da lam moi xong workbook: {result.workbook_path}")
    return 0


# ── Post-Write Sheet Formatters ──────────────────────────────────────────────
#
# Unified corporate palette and shared helpers used by ALL sheet formatters.
# Each sheet type (dashboard / analytics / flat-report) has its own function.
#

_CORP_COLORS: dict[str, Any] = {
    "title_bg":      "1E3A5F",   # Navy – title bars
    "title_fg":      "FFFFFF",
    "kpi_hdr_bg":    "1E40AF",   # Dark blue – KPI section header
    "kpi_hdr_fg":    "FFFFFF",
    "kpi_fgs":       ["2563EB", "16A34A", "7C3AED", "EA580C", "0891B2"],
    "kpi_bgs":       ["DBEAFE", "DCFCE7", "EDE9FE", "FEF3C7", "CFFAFE"],
    "section_bg":    "1D4ED8",   # Blue – section headers
    "section_fg":    "FFFFFF",
    "col_hdr_bg":    "DBEAFE",   # Light blue – column headers
    "col_hdr_fg":    "1E3A5F",
    "row_odd":       "FFFFFF",
    "row_even":      "EFF6FF",   # Pale blue – alternating rows
    "border":        "CBD5E1",
    "status_ok":     "DCFCE7",   # Green bg for OK status
    "status_warn":   "FEF3C7",   # Yellow bg for warning
    "status_err":    "FEE2E2",   # Red bg for error
}


# ── Shared style helpers (module-level, reused by every formatter) ────────────

def _fmt_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color.lstrip("#"))


def _fmt_font(bold: bool = False, size: int = 10, color: str = "000000") -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color.lstrip("#"))


def _fmt_align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _fmt_border(color: str = "CBD5E1", style: str = "thin") -> Border:
    s = Side(style=style, color=color.lstrip("#"))
    return Border(left=s, right=s, top=s, bottom=s)


def _fmt_norm(value: Any) -> str:
    """Strip Vietnamese diacritics and normalise for keyword matching."""
    if not value:
        return ""
    text = str(value).strip().upper()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return text


def _fmt_fit_columns(worksheet: Worksheet, min_w: int = 8, max_w: int = 40) -> None:
    """Auto-fit column widths based on content length, clamped to [min_w, max_w]."""
    col_max: dict[int, int] = {}
    for row_cells in worksheet.iter_rows():
        for cell in row_cells:
            if cell.value is not None:
                col_max[cell.column] = max(col_max.get(cell.column, 0), len(str(cell.value)))
    for col_idx, length in col_max.items():
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, min_w), max_w)


def _fmt_finish(worksheet: Worksheet, freeze: str = "A2") -> None:
    """Freeze pane and hide gridlines – called at the end of every formatter."""
    worksheet.freeze_panes = freeze
    worksheet.sheet_view.showGridLines = False


def _fmt_number(cell: Any) -> None:
    """Apply number format based on Python type."""
    if isinstance(cell.value, float):
        cell.number_format = "#,##0.00"
        cell.alignment = _fmt_align(h="right")
    elif isinstance(cell.value, int):
        cell.number_format = "#,##0"
        cell.alignment = _fmt_align(h="right")


# ── Dispatcher ───────────────────────────────────────────────────────────────

def _apply_post_write_formatting(sheet_name: str, worksheet: Worksheet) -> None:
    """Dispatch post-write formatting based on sheet name."""
    try:
        if sheet_name == SHEET_NAMES.get("analytics_grade_level"):
            _format_grade_level_sheet(worksheet)
        elif sheet_name in (
            SHEET_NAMES.get("dashboard_exec"),
            SHEET_NAMES.get("dashboard_operations"),
        ):
            _format_dashboard_sheet(worksheet)
        elif sheet_name in (
            SHEET_NAMES.get("analytics_course"),
            SHEET_NAMES.get("analytics_department"),
            SHEET_NAMES.get("analytics_department_course"),
            SHEET_NAMES.get("analytics_trend"),
        ):
            _format_analytics_sheet(worksheet)
        elif sheet_name in (
            SHEET_NAMES.get("report_external_assignment"),
            SHEET_NAMES.get("report_session_reconciliation"),
            SHEET_NAMES.get("looker_studio_data"),
        ):
            _format_flat_report_sheet(worksheet)
    except Exception:
        pass  # Formatting is cosmetic – never crash the main write flow


# ═══════════════════════════════════════════════════════════════════════════════
# FORMATTER A: Dashboard KPI sheets
# ═══════════════════════════════════════════════════════════════════════════════
#   Applies to: "Bảng điều khiển", "Điều hành hệ thống"
#   Layout:  Row 1 = title bar
#            Remaining rows = metric key-value pairs
#            Detect "metric_key" header → format as table with KPI highlights
# ───────────────────────────────────────────────────────────────────────────────

def _format_dashboard_sheet(worksheet: Worksheet) -> None:
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 6
    if max_row == 0:
        return

    C = _CORP_COLORS
    all_rows: list[tuple] = list(worksheet.iter_rows())

    # Classify: row 1 → title or header, rest → data
    for row_1idx, row_cells in enumerate(all_rows, start=1):
        first_val = ""
        for c in row_cells:
            if c.value is not None and str(c.value).strip():
                first_val = str(c.value).strip()
                break
        norm = _fmt_norm(first_val)

        if row_1idx == 1:
            # Check if it's a title or a header row
            is_header = any(
                kw in norm
                for kw in ("METRIC", "CHI_SO", "TEN_CHI_SO", "METRIC_KEY")
            )
            if is_header:
                # Format as column header
                for cell in row_cells:
                    cell.fill = _fmt_fill(C["col_hdr_bg"])
                    cell.font = _fmt_font(bold=True, size=10, color=C["col_hdr_fg"])
                    cell.alignment = _fmt_align(h="center", v="center", wrap=True)
                    cell.border = _fmt_border()
                worksheet.row_dimensions[row_1idx].height = 28
            else:
                # Format as title bar
                for cell in row_cells:
                    cell.fill = _fmt_fill(C["title_bg"])
                    cell.font = _fmt_font(bold=True, size=13, color=C["title_fg"])
                    cell.alignment = _fmt_align(h="left", v="center")
                worksheet.row_dimensions[row_1idx].height = 32
        else:
            # Data rows – alternating + KPI value highlighting
            bg = C["row_odd"] if (row_1idx % 2 == 0) else C["row_even"]
            for col_pos, cell in enumerate(row_cells):
                if cell.value is None:
                    continue
                cell.fill = _fmt_fill(bg)
                cell.border = _fmt_border()

                # Column 3 (metric_value) = big KPI number
                if col_pos == 2:
                    ci = (row_1idx - 2) % len(C["kpi_fgs"])
                    cell.font = _fmt_font(bold=True, size=14, color=C["kpi_fgs"][ci])
                    _fmt_number(cell)
                # Column 5 (status) = conditional color
                elif col_pos == 4:
                    status_norm = _fmt_norm(cell.value)
                    if "OK" in status_norm or "DAT" in status_norm:
                        cell.fill = _fmt_fill(C["status_ok"])
                    elif "WARN" in status_norm or "CANH BAO" in status_norm:
                        cell.fill = _fmt_fill(C["status_warn"])
                    elif "ERR" in status_norm or "LOI" in status_norm:
                        cell.fill = _fmt_fill(C["status_err"])
                    cell.font = _fmt_font(bold=True, size=9)
                    cell.alignment = _fmt_align(h="center")
                else:
                    cell.font = _fmt_font(size=9)
                    _fmt_number(cell)
                    if not isinstance(cell.value, (int, float)):
                        cell.alignment = _fmt_align(h="left", wrap=True)
            worksheet.row_dimensions[row_1idx].height = 22

    # AutoFilter on header (row 1 if header, row 2 if title+header)
    if max_row > 1 and max_col > 0:
        end_col = get_column_letter(min(max_col, 10))
        worksheet.auto_filter.ref = f"A1:{end_col}{max_row}"

    _fmt_fit_columns(worksheet)
    _fmt_finish(worksheet)


# ═══════════════════════════════════════════════════════════════════════════════
# FORMATTER B: Analytics multi-section sheets
# ═══════════════════════════════════════════════════════════════════════════════
#   Applies to: "Phân tích khóa học", "Phân tích phòng ban",
#               "Phòng ban theo khóa", "Phân tích xu hướng"
#   Layout:  Same as grade-level: title → KPI area → section headers → tables
#   Strategy: Generic 2-pass classifier (same as _format_grade_level_sheet
#             but with broader keyword detection)
# ───────────────────────────────────────────────────────────────────────────────

# Keywords that identify section headers across all analytics sheets
_ANALYTICS_SECTION_KEYWORDS = [
    "TONG HOP",
    "CHI TIET",
    "DANH SACH",
    "PHAN BO",
    "PHAN TICH",
    "XU HUONG",
    "THEO THANG",
    "THEO QUY",
]

# Keywords that identify column header rows
_ANALYTICS_COL_HDR_KEYWORDS = {
    "NGACH", "CAP BAC", "PHONG BAN", "HO TEN", "MA NV",
    "SO NHAN SU", "TEN KHOA HOC", "MA KHOA HOC", "NHOM KHOA HOC",
    "SO LUOT", "DIEM TB", "TY LE", "TONG GIO", "TONG CHI PHI",
    "THANG", "QUY", "NAM", "SO GIO", "SO HOC VIEN",
    "KHOA HOC", "COURSE", "DEPARTMENT", "TRAINING",
    "SO PHIEN", "DIEM HAI LONG", "DIEM NPS",
}


def _format_analytics_sheet(worksheet: Worksheet) -> None:  # noqa: C901
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 12
    if max_row == 0:
        return

    C = _CORP_COLORS

    def _first_nonempty(row_cells: tuple) -> str:
        for cell in row_cells:
            if cell.value is not None and str(cell.value).strip():
                return str(cell.value).strip()
        return ""

    def _nonempty_count(row_cells: tuple) -> int:
        return sum(1 for c in row_cells if c.value is not None and str(c.value).strip())

    def _is_section_hdr(norm_val: str) -> bool:
        return any(kw.replace(" ", "") in norm_val.replace(" ", "") for kw in _ANALYTICS_SECTION_KEYWORDS)

    def _is_col_hdr(norm_val: str) -> bool:
        tokens = norm_val.replace("_", " ").split()
        return any(kw in tokens or kw in norm_val for kw in _ANALYTICS_COL_HDR_KEYWORDS)

    # ── Pass 1: classify ─────────────────────────────────────────────────────
    all_rows: list[tuple] = list(worksheet.iter_rows())
    row_types: list[str] = []
    first_section_found = False

    for row_cells in all_rows:
        first = _first_nonempty(row_cells)
        norm = _fmt_norm(first)
        n = _nonempty_count(row_cells)

        if n == 0:
            row_types.append("empty")
        elif ("BAO CAO" in norm or "PHAN TICH" in norm) and not first_section_found and n <= 3:
            row_types.append("title")
        elif not first_section_found and norm.strip() == "KPI":
            row_types.append("kpi_section_hdr")
        elif _is_section_hdr(norm) and n <= 4:
            first_section_found = True
            row_types.append("section_hdr")
        elif _is_col_hdr(norm) and n >= 2:
            if not first_section_found:
                first_section_found = True
            row_types.append("col_header")
        elif not first_section_found:
            has_numeric = any(isinstance(c.value, (int, float)) for c in row_cells if c.value is not None)
            row_types.append("kpi_value" if has_numeric else "kpi_label")
        else:
            row_types.append("data")

    # ── Pass 2: style ────────────────────────────────────────────────────────
    kpi_idx = 0
    data_row_count = 0
    col_hdr_row = None
    last_data_row = None
    prev_ncols = 0
    af_ranges: list[tuple[int, int, int]] = []

    for r1, (row_cells, rtype) in enumerate(zip(all_rows, row_types), start=1):
        if rtype == "title":
            for cell in row_cells:
                cell.fill = _fmt_fill(C["title_bg"])
                cell.font = _fmt_font(bold=True, size=13, color=C["title_fg"])
                cell.alignment = _fmt_align(h="left", v="center")
            worksheet.row_dimensions[r1].height = 32

        elif rtype == "kpi_section_hdr":
            for cell in row_cells:
                cell.fill = _fmt_fill(C["kpi_hdr_bg"])
                cell.font = _fmt_font(bold=True, size=10, color=C["kpi_hdr_fg"])
                cell.alignment = _fmt_align(h="left")
            worksheet.row_dimensions[r1].height = 18

        elif rtype == "kpi_label":
            positions = [j for j, c in enumerate(row_cells) if c.value is not None and str(c.value).strip()]
            for si, cp in enumerate(positions):
                ci = (kpi_idx + si) % len(C["kpi_fgs"])
                row_cells[cp].fill = _fmt_fill(C["kpi_fgs"][ci])
                row_cells[cp].font = _fmt_font(bold=True, size=9, color="FFFFFF")
                row_cells[cp].alignment = _fmt_align(h="left", v="bottom")
                nxt = positions[si + 1] if si + 1 < len(positions) else cp + 3
                for fp in range(cp + 1, min(nxt, len(row_cells))):
                    row_cells[fp].fill = _fmt_fill(C["kpi_fgs"][ci])
            kpi_idx += len(positions)
            worksheet.row_dimensions[r1].height = 22

        elif rtype == "kpi_value":
            positions = [j for j, c in enumerate(row_cells) if c.value is not None and str(c.value).strip()]
            start = (kpi_idx - len(positions)) % len(C["kpi_bgs"])
            for si, cp in enumerate(positions):
                ci = (start + si) % len(C["kpi_bgs"])
                row_cells[cp].fill = _fmt_fill(C["kpi_bgs"][ci])
                row_cells[cp].font = _fmt_font(bold=True, size=16, color=C["kpi_fgs"][ci])
                row_cells[cp].alignment = _fmt_align(h="left", v="top")
                _fmt_number(row_cells[cp])
                nxt = positions[si + 1] if si + 1 < len(positions) else cp + 3
                for fp in range(cp + 1, min(nxt, len(row_cells))):
                    row_cells[fp].fill = _fmt_fill(C["kpi_bgs"][ci])
            worksheet.row_dimensions[r1].height = 30

        elif rtype == "section_hdr":
            if col_hdr_row is not None and last_data_row is not None:
                af_ranges.append((col_hdr_row, last_data_row, prev_ncols))
            col_hdr_row = None
            last_data_row = None
            data_row_count = 0
            for cell in row_cells:
                cell.fill = _fmt_fill(C["section_bg"])
                cell.font = _fmt_font(bold=True, size=11, color=C["section_fg"])
                cell.alignment = _fmt_align(h="left")
            worksheet.row_dimensions[r1].height = 24

        elif rtype == "col_header":
            prev_ncols = _nonempty_count(row_cells)
            col_hdr_row = r1
            data_row_count = 0
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = _fmt_fill(C["col_hdr_bg"])
                    cell.font = _fmt_font(bold=True, size=9, color=C["col_hdr_fg"])
                    cell.alignment = _fmt_align(h="center", v="center", wrap=True)
                    cell.border = _fmt_border()
            worksheet.row_dimensions[r1].height = 30

        elif rtype == "data":
            last_data_row = r1
            bg = C["row_odd"] if data_row_count % 2 == 0 else C["row_even"]
            data_row_count += 1
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = _fmt_fill(bg)
                    cell.font = _fmt_font(size=9)
                    cell.border = _fmt_border()
                    _fmt_number(cell)
                    if not isinstance(cell.value, (int, float)):
                        cell.alignment = _fmt_align(h="left", wrap=True)
            worksheet.row_dimensions[r1].height = 16

        elif rtype == "empty":
            worksheet.row_dimensions[r1].height = 6

    # Close last section
    if col_hdr_row is not None and last_data_row is not None:
        af_ranges.append((col_hdr_row, last_data_row, prev_ncols))

    # AutoFilter on the last (most detailed) table
    if af_ranges:
        hdr, last, ncols = af_ranges[-1]
        if last > hdr and ncols > 0:
            worksheet.auto_filter.ref = f"A{hdr}:{get_column_letter(min(ncols, max_col))}{last}"

    _fmt_fit_columns(worksheet)
    _fmt_finish(worksheet)


# ═══════════════════════════════════════════════════════════════════════════════
# FORMATTER C: Flat report sheets
# ═══════════════════════════════════════════════════════════════════════════════
#   Applies to: "Cử đi học bên ngoài", "Đối chiếu lớp đào tạo"
#   Layout:  Row 1 = header (or title → row 2 = header)
#            Remaining rows = data
# ───────────────────────────────────────────────────────────────────────────────

def _format_flat_report_sheet(worksheet: Worksheet) -> None:
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 8
    if max_row == 0:
        return

    C = _CORP_COLORS
    all_rows: list[tuple] = list(worksheet.iter_rows())

    # Detect: is row 1 a title or a column header?
    first_norm = _fmt_norm(all_rows[0][0].value if all_rows[0][0].value else "") if all_rows else ""
    has_title = any(kw in first_norm for kw in ("BAO CAO", "CU DI HOC", "DOI CHIEU", "DANH SACH"))

    header_row_idx = 2 if has_title else 1  # 1-indexed
    data_start = header_row_idx + 1
    data_count = 0

    for r1, row_cells in enumerate(all_rows, start=1):
        # Title row
        if has_title and r1 == 1:
            for cell in row_cells:
                cell.fill = _fmt_fill(C["title_bg"])
                cell.font = _fmt_font(bold=True, size=13, color=C["title_fg"])
                cell.alignment = _fmt_align(h="left", v="center")
            worksheet.row_dimensions[r1].height = 32

        # Header row
        elif r1 == header_row_idx:
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = _fmt_fill(C["col_hdr_bg"])
                    cell.font = _fmt_font(bold=True, size=10, color=C["col_hdr_fg"])
                    cell.alignment = _fmt_align(h="center", v="center", wrap=True)
                    cell.border = _fmt_border()
            worksheet.row_dimensions[r1].height = 28

        # Data rows
        elif r1 >= data_start:
            nonempty = sum(1 for c in row_cells if c.value is not None and str(c.value).strip())
            if nonempty == 0:
                worksheet.row_dimensions[r1].height = 6
                continue
            bg = C["row_odd"] if data_count % 2 == 0 else C["row_even"]
            data_count += 1
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = _fmt_fill(bg)
                    cell.font = _fmt_font(size=9)
                    cell.border = _fmt_border()
                    _fmt_number(cell)
                    if not isinstance(cell.value, (int, float)):
                        cell.alignment = _fmt_align(h="left", wrap=True)
            worksheet.row_dimensions[r1].height = 18

    # AutoFilter on header row
    if max_row > header_row_idx and max_col > 0:
        end_col = get_column_letter(min(max_col, 20))
        worksheet.auto_filter.ref = f"A{header_row_idx}:{end_col}{max_row}"

    _fmt_fit_columns(worksheet)
    freeze_cell = f"A{header_row_idx + 1}"
    _fmt_finish(worksheet, freeze=freeze_cell)


# ═══════════════════════════════════════════════════════════════════════════════
# FORMATTER D: Grade-Level Analytics sheet (original, updated to use shared helpers)
# ═══════════════════════════════════════════════════════════════════════════════

# Grade-level specific keywords
_GL_SECTION_KEYWORDS = [
    "TONG HOP THEO NGACH",
    "TONG HOP THEO CAP BAC",
    "CHI TIET PHONG BAN",
    "DANH SACH HOC VIEN",
]
_GL_COL_HDR_KEYWORDS = {
    "NGACH", "CAP BAC", "PHONG BAN", "HO TEN", "MA NV",
    "SO NHAN SU", "CUP BAC",
}


def _format_grade_level_sheet(worksheet: Worksheet) -> None:  # noqa: C901
    """Apply professional Excel formatting to the grade-level analytics sheet.

    Strategy:
      Pass 1 – scan every row and classify its role (title / kpi_section_hdr /
               kpi_label / kpi_value / section_hdr / col_header / data / empty).
      Pass 2 – apply openpyxl styles, row heights, column widths, AutoFilter,
               freeze pane, and hide grid lines.
    """
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 12
    if max_row == 0:
        return

    C = _CORP_COLORS

    def _first_nonempty(row_cells: tuple) -> str:
        for cell in row_cells:
            if cell.value is not None and str(cell.value).strip():
                return str(cell.value).strip()
        return ""

    def _nonempty_count(row_cells: tuple) -> int:
        return sum(1 for c in row_cells if c.value is not None and str(c.value).strip())

    def _is_section_hdr(norm_val: str) -> bool:
        return any(kw.replace(" ", "") in norm_val.replace(" ", "") for kw in _GL_SECTION_KEYWORDS)

    def _is_col_hdr(norm_val: str) -> bool:
        return any(kw in norm_val for kw in _GL_COL_HDR_KEYWORDS)

    # ── Pass 1: classify each row ─────────────────────────────────────────────

    all_rows: list[tuple] = list(worksheet.iter_rows())
    row_types: list[str] = []
    first_section_found = False

    for row_cells in all_rows:
        first = _first_nonempty(row_cells)
        norm  = _fmt_norm(first)
        n_nonempty = _nonempty_count(row_cells)

        if n_nonempty == 0:
            row_types.append("empty")
        elif "BAO CAO PHAN TICH" in norm or ("BAO CAO" in norm and not first_section_found):
            row_types.append("title")
        elif not first_section_found and norm.strip() == "KPI":
            row_types.append("kpi_section_hdr")
        elif _is_section_hdr(norm):
            first_section_found = True
            row_types.append("section_hdr")
        elif _is_col_hdr(norm) and n_nonempty >= 2:
            row_types.append("col_header")
        elif not first_section_found:
            has_numeric = any(isinstance(c.value, (int, float)) for c in row_cells if c.value is not None)
            row_types.append("kpi_value" if has_numeric else "kpi_label")
        else:
            row_types.append("data")

    # ── Pass 2: apply formatting ──────────────────────────────────────────────

    kpi_item_global_idx = 0
    section_idx          = -1
    data_row_in_section  = 0
    prev_col_hdr_ncols   = 0
    col_hdr_row_1idx     = None
    last_data_row_1idx   = None
    autofilter_ranges: list[tuple[int, int, int]] = []

    for row_1idx, (row_cells, row_type) in enumerate(zip(all_rows, row_types), start=1):

        if row_type == "title":
            for cell in row_cells:
                cell.fill = _fmt_fill(C["title_bg"])
                cell.font = _fmt_font(bold=True, size=13, color=C["title_fg"])
                has_value = cell.value is not None and str(cell.value).strip()
                is_primary = has_value and "BAO CAO" in _fmt_norm(cell.value)
                cell.alignment = _fmt_align(h="left" if is_primary else "right", v="center")
            worksheet.row_dimensions[row_1idx].height = 32

        elif row_type == "kpi_section_hdr":
            for cell in row_cells:
                cell.fill = _fmt_fill(C["kpi_hdr_bg"])
                cell.font = _fmt_font(bold=True, size=10, color=C["kpi_hdr_fg"])
                cell.alignment = _fmt_align(h="left", v="center")
            worksheet.row_dimensions[row_1idx].height = 18

        elif row_type == "kpi_label":
            nonempty_positions = [j for j, c in enumerate(row_cells) if c.value is not None and str(c.value).strip()]
            for slot_idx, col_pos in enumerate(nonempty_positions):
                ci = (kpi_item_global_idx + slot_idx) % len(C["kpi_fgs"])
                cell = row_cells[col_pos]
                cell.fill = _fmt_fill(C["kpi_fgs"][ci])
                cell.font = _fmt_font(bold=True, size=9, color="FFFFFF")
                cell.alignment = _fmt_align(h="left", v="bottom")
                next_p = nonempty_positions[slot_idx + 1] if slot_idx + 1 < len(nonempty_positions) else col_pos + 3
                for fp in range(col_pos + 1, min(next_p, len(row_cells))):
                    row_cells[fp].fill = _fmt_fill(C["kpi_fgs"][ci])
            kpi_item_global_idx += len(nonempty_positions)
            worksheet.row_dimensions[row_1idx].height = 22

        elif row_type == "kpi_value":
            nonempty_positions = [j for j, c in enumerate(row_cells) if c.value is not None and str(c.value).strip()]
            start_ci = (kpi_item_global_idx - len(nonempty_positions)) % len(C["kpi_bgs"])
            for slot_idx, col_pos in enumerate(nonempty_positions):
                ci = (start_ci + slot_idx) % len(C["kpi_bgs"])
                cell = row_cells[col_pos]
                cell.fill = _fmt_fill(C["kpi_bgs"][ci])
                cell.font = _fmt_font(bold=True, size=16, color=C["kpi_fgs"][ci])
                cell.alignment = _fmt_align(h="left", v="top")
                _fmt_number(cell)
                next_p = nonempty_positions[slot_idx + 1] if slot_idx + 1 < len(nonempty_positions) else col_pos + 3
                for fp in range(col_pos + 1, min(next_p, len(row_cells))):
                    row_cells[fp].fill = _fmt_fill(C["kpi_bgs"][ci])
            worksheet.row_dimensions[row_1idx].height = 30

        elif row_type == "section_hdr":
            if col_hdr_row_1idx is not None and last_data_row_1idx is not None:
                autofilter_ranges.append((col_hdr_row_1idx, last_data_row_1idx, prev_col_hdr_ncols))
            col_hdr_row_1idx = None
            last_data_row_1idx = None
            section_idx += 1
            data_row_in_section = 0
            for cell in row_cells:
                cell.fill = _fmt_fill(C["section_bg"])
                cell.font = _fmt_font(bold=True, size=11, color=C["section_fg"])
                cell.alignment = _fmt_align(h="left", v="center")
            worksheet.row_dimensions[row_1idx].height = 24

        elif row_type == "col_header":
            prev_col_hdr_ncols = _nonempty_count(row_cells)
            col_hdr_row_1idx = row_1idx
            data_row_in_section = 0
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = _fmt_fill(C["col_hdr_bg"])
                    cell.font = _fmt_font(bold=True, size=9, color=C["col_hdr_fg"])
                    cell.alignment = _fmt_align(h="center", v="center", wrap=True)
                    cell.border = _fmt_border()
            worksheet.row_dimensions[row_1idx].height = 30

        elif row_type == "data":
            last_data_row_1idx = row_1idx
            bg = C["row_odd"] if data_row_in_section % 2 == 0 else C["row_even"]
            data_row_in_section += 1
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip():
                    cell.fill = _fmt_fill(bg)
                    cell.font = _fmt_font(size=9)
                    cell.border = _fmt_border()
                    _fmt_number(cell)
                    if not isinstance(cell.value, (int, float)):
                        cell.alignment = _fmt_align(h="left", wrap=True)
            worksheet.row_dimensions[row_1idx].height = 16

        elif row_type == "empty":
            worksheet.row_dimensions[row_1idx].height = 6

    # Close last section
    if col_hdr_row_1idx is not None and last_data_row_1idx is not None:
        autofilter_ranges.append((col_hdr_row_1idx, last_data_row_1idx, prev_col_hdr_ncols))

    # AutoFilter on the last (most detailed) data table
    if autofilter_ranges:
        hdr_row, last_data, n_cols = autofilter_ranges[-1]
        if last_data > hdr_row and n_cols > 0:
            worksheet.auto_filter.ref = f"A{hdr_row}:{get_column_letter(min(n_cols, max_col))}{last_data}"

    _fmt_fit_columns(worksheet)
    _fmt_finish(worksheet)


# ── Entry Point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    raise SystemExit(main())
