"""HR-NEXUS Google Sheets ↔ Local Python Sync Bridge.

Transport layer only — all business logic delegates to existing modules:
  - HEADER_ALIASES → local_excel_runner.py
  - build_local_training_sync() → local_fact_builder.py
  - transform_data() → transform.py

Usage:
  python gsheet_sync.py check       # Kiểm tra kết nối + liệt kê sheets
  python gsheet_sync.py sync        # Full pipeline: pull → process → push + mirror
  python gsheet_sync.py pull-only   # Chỉ kéo data xuống file Excel local
  python gsheet_sync.py push-only   # Từ Excel local → đẩy kết quả lên Sheets
  python gsheet_sync.py push-full   # Đẩy toàn bộ Excel local ghi đè lên Sheets
"""
from __future__ import annotations

import argparse
import json
import sys
import time
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import gspread
import pandas as pd
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook as _openpyxl_load_workbook

from local_excel_runner import (
    DEFAULT_WORKBOOK_PATH,
    HEADER_ALIASES,
    SOURCE_SHEETS,
    REQUIRED_COLUMNS,
    fetch_data,
    write_data,
)
from local_fact_builder import build_local_training_sync
from transform import AnalyticsInputs, SHEET_NAMES, transform_data

Logger = Callable[[str], None]

# ── Config & Constants ───────────────────────────────────────────────────────

DEFAULT_CONFIG_PATH = Path(__file__).resolve().parent / "sync_config.json"

# Sheets to PULL from Google Sheets (source data)
PULL_SHEETS: tuple[str, ...] = SOURCE_SHEETS

# Sheets to PUSH to Google Sheets (report output)
PUSH_SHEET_NAMES: dict[str, str] = SHEET_NAMES

# Google API scopes
_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly",
]

# Sheet name constants for easy reference
from local_excel_runner import (
    EMPLOYEES_SHEET,
    COURSES_SHEET,
    TRAINING_SESSIONS_SHEET,
    RAW_PARTICIPANTS_SHEET,
    TRAINING_RECORDS_SHEET,
    QUEUE_JOBS_SHEET,
    QA_RESULTS_SHEET,
    STAGING_HR_SHEET,
    STAGING_COURSES_SHEET,
    STAGING_TRAINING_SHEET,
    CONFIG_SHEET,
)

# CRITICAL sheets that MUST exist for the engine to work
MANDATORY_SHEETS = (
    EMPLOYEES_SHEET,
    TRAINING_SESSIONS_SHEET,
    RAW_PARTICIPANTS_SHEET,
    CONFIG_SHEET,
)


# ── UX-Friendly Error Messages (SRS E01-E08) ────────────────────────────────

FRIENDLY_ERRORS: dict[str, str] = {
    "E01_SPAM_CLICK": (
        "Hệ thống đang xử lý dữ liệu cho bạn rồi, "
        "vui lòng đợi trong giây lát nhé!"
    ),
    "E02_TIMEOUT": (
        "Dữ liệu đợt này hơi lớn nên cần thêm chút thời gian. "
        "Hệ thống vẫn đang chạy ngầm an toàn nhé."
    ),
    "E03_BAD_DATA": (
        "Có vẻ một vài dòng dữ liệu đang bị sai định dạng ({details}). "
        "Bạn kiểm tra lại giúp hệ thống nhé!"
    ),
    "E04_NETWORK": (
        "Đường truyền kết nối đang bị gián đoạn. Không sao cả, "
        "dữ liệu của bạn vẫn an toàn. Hãy thử lại sau ít phút nhé."
    ),
    "E05_PERMISSION": (
        "Không ghi được vào sheet '{sheet}'. "
        "Bạn kiểm tra quyền chỉnh sửa trên Google Sheets giúp nhé."
    ),
    "E06_NO_CREDS": (
        "Chưa có file xác thực Google. "
        "Hãy tạo Service Account theo docs/CREDENTIALS_SETUP.md"
    ),
    "E07_RATE_LIMIT": (
        "Google đang bận, hệ thống tự thử lại... (lần {attempt}/3)"
    ),
    "E08_SCHEMA_CHANGE": (
        "Sheet '{sheet}' thiếu cột '{column}'. "
        "Dữ liệu vẫn xử lý được nhưng thiếu một số phân tích."
    ),
}


class SyncError(Exception):
    """User-friendly sync error with error code."""

    def __init__(self, code: str, message: str, cause: Exception | None = None):
        self.code = code
        self.friendly_message = message
        self.__cause__ = cause
        super().__init__(message)


# ── Data Classes ─────────────────────────────────────────────────────────────


@dataclass(frozen=True)
class SyncConfig:
    spreadsheet_id: str
    credentials_path: Path
    fiscal_year: str
    backup_before_push: bool = True
    gcp_project_id: str = ""

    @staticmethod
    def load(config_path: str | Path = DEFAULT_CONFIG_PATH) -> SyncConfig:
        path = Path(config_path)
        if not path.exists():
            raise SyncError(
                "E06_NO_CREDS",
                f"Không tìm thấy file cấu hình: {path}. "
                "Hãy tạo sync_config.json theo hướng dẫn.",
            )
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        return SyncConfig(
            spreadsheet_id=data["spreadsheet_id"],
            credentials_path=Path(data.get("credentials_path", "credentials/service_account.json")),
            fiscal_year=str(data.get("fiscal_year", datetime.now().year)),
            backup_before_push=data.get("backup_before_push", True),
            gcp_project_id=data.get("gcp_project_id", ""),
        )


@dataclass
class SyncResult:
    success: bool = True
    sheets_pulled: int = 0
    sheets_pushed: int = 0
    total_rows_pulled: int = 0
    total_rows_pushed: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    duration_seconds: float = 0.0
    mirror_path: Path | None = None


# ── Connection ───────────────────────────────────────────────────────────────


def connect(config: SyncConfig, logger: Logger | None = None) -> gspread.Spreadsheet:
    """Connect to Google Sheets via Service Account. (E04, E06)"""
    creds_path = config.credentials_path
    if not creds_path.is_absolute():
        creds_path = Path(__file__).resolve().parent / creds_path

    # E06: Pre-flight credential check
    if not creds_path.exists():
        raise SyncError(
            "E06_NO_CREDS",
            FRIENDLY_ERRORS["E06_NO_CREDS"],
        )

    try:
        _emit(logger, "Đang kết nối Google Sheets...")
        credentials = Credentials.from_service_account_file(str(creds_path), scopes=_SCOPES)
        client = gspread.authorize(credentials)
        spreadsheet = client.open_by_key(config.spreadsheet_id)
        _emit(logger, f"✅ Đã kết nối: {spreadsheet.title}")
        return spreadsheet
    except gspread.exceptions.SpreadsheetNotFound:
        raise SyncError(
            "E05_PERMISSION",
            FRIENDLY_ERRORS["E05_PERMISSION"].format(sheet="toàn bộ spreadsheet"),
            cause=None,
        )
    except (ConnectionError, TimeoutError, OSError) as exc:
        raise SyncError("E04_NETWORK", FRIENDLY_ERRORS["E04_NETWORK"], cause=exc)
    except Exception as exc:
        raise SyncError("E04_NETWORK", FRIENDLY_ERRORS["E04_NETWORK"], cause=exc)


# ── PULL: Google Sheets → DataFrame ─────────────────────────────────────────


def _normalize_token_simple(value: Any) -> str:
    """Simplified normalize for header matching (mirrors local_excel_runner)."""
    import unicodedata

    text = str(value or "").strip()
    text = text.replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii").lower()
    chars = [c if c.isalnum() else "_" for c in text]
    normalized = "".join(chars).strip("_")
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized


def pull_sheet(
    spreadsheet: gspread.Spreadsheet,
    sheet_name: str,
    logger: Logger | None = None,
) -> pd.DataFrame:
    """Pull 1 sheet from Google Sheets → DataFrame with alias mapping. (E05, E08)"""
    # Find sheet by normalized name
    worksheet = _find_worksheet(spreadsheet, sheet_name)
    if worksheet is None:
        _emit(logger, f"  ⚠️ Sheet '{sheet_name}' không tìm thấy, trả về DataFrame rỗng.")
        columns = REQUIRED_COLUMNS.get(sheet_name, ())
        return pd.DataFrame({col: pd.Series(dtype="object") for col in columns})

    try:
        all_values = worksheet.get_all_values()
    except gspread.exceptions.APIError as exc:
        if exc.response.status_code == 403:
            raise SyncError(
                "E05_PERMISSION",
                FRIENDLY_ERRORS["E05_PERMISSION"].format(sheet=sheet_name),
                cause=exc,
            )
        raise

    if not all_values:
        return pd.DataFrame()

    headers = [str(h).strip() for h in all_values[0]]
    body = all_values[1:]

    # Build DataFrame
    if not body:
        frame = pd.DataFrame(columns=headers)
    else:
        frame = pd.DataFrame(body, columns=headers)
        # Replace empty strings with None for consistency
        frame = frame.replace("", None)

    # Apply HEADER_ALIASES (same logic as local_excel_runner._load_sheet)
    alias_map = HEADER_ALIASES.get(sheet_name, {})
    rename_map = {
        col: alias_map[_normalize_token_simple(col)]
        for col in frame.columns
        if _normalize_token_simple(col) in alias_map
    }
    frame = frame.rename(columns=rename_map)

    # Add __rowNumber for compatibility
    frame["__rowNumber"] = pd.Series(range(2, len(frame) + 2), dtype="Int64")

    # E08: Schema validation — warn about missing required columns
    required = REQUIRED_COLUMNS.get(sheet_name, ())
    missing = [col for col in required if col not in frame.columns]
    if missing:
        for col in missing:
            _emit(
                logger,
                f"  ⚠️ {FRIENDLY_ERRORS['E08_SCHEMA_CHANGE'].format(sheet=sheet_name, column=col)}",
            )
            frame[col] = None  # Add missing column with None values

    row_count = len(frame)
    _emit(logger, f"  📥 {sheet_name}: {row_count} dòng, {len(rename_map)} cột đã map")
    return frame


def pull_all_sources(
    spreadsheet: gspread.Spreadsheet,
    logger: Logger | None = None,
) -> dict[str, pd.DataFrame]:
    """Pull all 11 source sheets → dict of DataFrames. (E07 retry)"""
    _emit(logger, "─── PULL: Đang kéo dữ liệu từ Google Sheets ───")
    tables: dict[str, pd.DataFrame] = {}
    
    # Check mandatory sheets first
    missing_mandatory = []
    for sheet_name in MANDATORY_SHEETS:
        if _find_worksheet(spreadsheet, sheet_name) is None:
            missing_mandatory.append(sheet_name)
    
    if missing_mandatory:
        raise SyncError(
            "E08_SCHEMA_CHANGE",
            f"Thieu cac sheet bat buoc: {', '.join(missing_mandatory)}. "
            "Anh hay kiem tra lai file Google Sheet nhe!"
        )

    for sheet_name in PULL_SHEETS:
        tables[sheet_name] = _retry_with_backoff(
            lambda sn=sheet_name: pull_sheet(spreadsheet, sn, logger),
            logger=logger,
        )
    total = sum(len(df) for df in tables.values())
    _emit(logger, f"✅ Đã kéo xong {len(tables)} sheets, tổng {total} dòng")
    return tables


# ── PROCESS: Reuse existing engine ──────────────────────────────────────────


def process(
    tables: dict[str, pd.DataFrame],
    fiscal_year: str,
    logger: Logger | None = None,
) -> tuple[dict[str, list[list[Any]]], dict[str, pd.DataFrame]]:
    """Run sync engine + transform. Returns (outputs, source_updates). (E03)"""
    _emit(logger, "─── PROCESS: Đang xử lý dữ liệu ───")

    # Import sheet name constants
    from local_excel_runner import (
        EMPLOYEES_SHEET,
        COURSES_SHEET,
        TRAINING_SESSIONS_SHEET,
        RAW_PARTICIPANTS_SHEET,
        TRAINING_RECORDS_SHEET,
        QUEUE_JOBS_SHEET,
        QA_RESULTS_SHEET,
        STAGING_HR_SHEET,
        STAGING_COURSES_SHEET,
        STAGING_TRAINING_SHEET,
        CONFIG_SHEET,
    )

    try:
        # Step 1: Build local training sync (same as run_local_refresh)
        _emit(logger, "  🔄 Đang đồng bộ fact table...")
        local_sync = build_local_training_sync(
            employees=tables[EMPLOYEES_SHEET],
            courses=tables[COURSES_SHEET],
            training_sessions=tables[TRAINING_SESSIONS_SHEET],
            raw_participants=tables[RAW_PARTICIPANTS_SHEET],
            existing_records=tables[TRAINING_RECORDS_SHEET],
            logger=lambda msg: _emit(logger, f"    {msg}"),
        )

        # Step 2: Build AnalyticsInputs (same as run_local_refresh)
        from local_excel_runner import _build_config_map, _build_raw_sync_summary

        config_map = _build_config_map(tables[CONFIG_SHEET])
        raw_sync_summary = _build_raw_sync_summary(tables[QUEUE_JOBS_SHEET])
        raw_sync_summary["failed_rows"] = local_sync.failed_rows

        inputs = AnalyticsInputs(
            training_records=local_sync.training_records,
            employees=tables[EMPLOYEES_SHEET],
            training_sessions=tables[TRAINING_SESSIONS_SHEET],
            raw_participants=local_sync.raw_participants,
            queue_jobs=tables[QUEUE_JOBS_SHEET],
            qa_results=tables[QA_RESULTS_SHEET],
            staging_hr=tables[STAGING_HR_SHEET],
            staging_courses=tables[STAGING_COURSES_SHEET],
            staging_training=tables[STAGING_TRAINING_SHEET],
            config_map=config_map,
            raw_sync_summary=raw_sync_summary,
        )

        # Step 3: Transform
        _emit(logger, "  📊 Đang tính toán 10 sheet báo cáo bằng pandas...")
        outputs = transform_data(inputs, fiscal_year=fiscal_year)

        source_updates = {
            RAW_PARTICIPANTS_SHEET: local_sync.raw_participants,
            TRAINING_RECORDS_SHEET: local_sync.training_records,
            EMPLOYEES_SHEET: local_sync.employees,
        }

        total_rows = sum(len(matrix) for matrix in outputs.values())
        _emit(logger, f"  ✅ Đã tạo {len(outputs)} sheets, tổng {total_rows} dòng")
        return outputs, source_updates

    except Exception as exc:
        raise SyncError(
            "E03_BAD_DATA",
            FRIENDLY_ERRORS["E03_BAD_DATA"].format(details=str(exc)),
            cause=exc,
        )


# ── PUSH: DataFrame → Google Sheets ─────────────────────────────────────────


def push_sheet(
    spreadsheet: gspread.Spreadsheet,
    sheet_name: str,
    matrix: list[list[Any]],
    logger: Logger | None = None,
) -> int:
    """Push 1 matrix to a Google Sheet (clear + batch update). (E05, E07)"""
    worksheet = _find_worksheet(spreadsheet, sheet_name)
    if worksheet is None:
        try:
            worksheet = spreadsheet.add_worksheet(sheet_name, rows=len(matrix) + 10, cols=20)
        except Exception:
            _emit(logger, f"  ⚠️ Không tạo được sheet '{sheet_name}', bỏ qua.")
            return 0

    if logger:
        _emit(logger, f"  📤 [GID:{worksheet.id}] ?ang ghi '{sheet_name}'...")

    try:
        # Clear existing data but preserve formatting
        worksheet.batch_clear(["A1:ZZ10000"])
        # Batch update all cells
        if matrix:
            safe_matrix = _sheets_safe_matrix(matrix)
            worksheet.update(safe_matrix, value_input_option="USER_ENTERED")
        row_count = len(matrix)
        _emit(logger, f"  ✅ {sheet_name}: {row_count} d?ng ?? ghi (GID:{worksheet.id})")
        return row_count
    except gspread.exceptions.APIError as exc:
        if exc.response.status_code == 403:
            raise SyncError(
                "E05_PERMISSION",
                FRIENDLY_ERRORS["E05_PERMISSION"].format(sheet=sheet_name),
                cause=exc,
            )
        raise


def push_all_outputs(
    spreadsheet: gspread.Spreadsheet,
    outputs: dict[str, list[list[Any]]],
    logger: Logger | None = None,
) -> int:
    """Push all output sheets to Google Sheets. (E07 retry per sheet)"""
    _emit(logger, "─── PUSH: Đang ghi kết quả lên Cloud Sheets ───")
    total_pushed = 0
    
    # Filter only output sheets if necessary, or push everything generated
    # (Checking if names match our expected output list to prevent garbage sheets)
    from local_excel_runner import OUTPUT_SHEETS
    valid_outputs = {k: v for k, v in outputs.items() if k in OUTPUT_SHEETS}
    
    if len(valid_outputs) < len(outputs):
        skipped = [k for k in outputs.keys() if k not in OUTPUT_SHEETS]
        _emit(logger, f"  ⚠️ Bỏ qua {len(skipped)} sheet không nằm trong OUTPUT_SHEETS")

    for sheet_name, matrix in valid_outputs.items():
        pushed = _retry_with_backoff(
            lambda sn=sheet_name, mx=matrix: push_sheet(spreadsheet, sn, mx, logger),
            logger=logger,
        )
        total_pushed += pushed
        time.sleep(0.5)
        
    _emit(logger, f"🚀 Hoàn tất: Đã cập nhật {len(valid_outputs)} sheets, tổng {total_pushed} dòng.")
    return total_pushed


# ── Mirror to Excel ─────────────────────────────────────────────────────────


def mirror_to_excel(
    outputs: dict[str, list[list[Any]]],
    source_updates: dict[str, pd.DataFrame] | None = None,
    workbook_path: str | Path = DEFAULT_WORKBOOK_PATH,
    backup: bool = True,
    logger: Logger | None = None,
) -> Path | None:
    """Mirror results to local Excel file (same as local runner)."""
    _emit(logger, "─── MIRROR: Cập nhật file Excel local ───")
    try:
        backup_path = write_data(
            workbook_path,
            outputs,
            source_updates=source_updates,
            backup=backup,
            logger=lambda msg: _emit(logger, f"  {msg}"),
        )
        _emit(logger, "✅ Đã cập nhật file Excel local")
        return backup_path
    except PermissionError:
        _emit(
            logger,
            "⚠️ Không ghi được file Excel (có thể đang mở). "
            "Kết quả trên Google Sheets vẫn OK.",
        )
        return None
    except Exception as exc:
        _emit(logger, f"⚠️ Lỗi khi ghi Excel local: {exc}. Kết quả trên Sheets vẫn OK.")
        return None


# ── Orchestrators ────────────────────────────────────────────────────────────


def run_check(config_path: str | Path = DEFAULT_CONFIG_PATH, logger: Logger | None = None) -> None:
    """Check connection and list sheets."""
    config = SyncConfig.load(config_path)
    spreadsheet = connect(config, logger)
    _emit(logger, f"\n📋 Spreadsheet: {spreadsheet.title}")
    _emit(logger, f"   ID: {config.spreadsheet_id}")
    _emit(logger, f"   GCP Project: {config.gcp_project_id}\n")

    worksheets = spreadsheet.worksheets()
    _emit(logger, f"📊 Tìm thấy {len(worksheets)} sheets:")
    for ws in worksheets:
        _emit(logger, f"   • {ws.title} ({ws.row_count} dòng × {ws.col_count} cột)")

    # Check source sheets availability
    _emit(logger, "\n🔍 Kiểm tra 11 sheet nguồn:")
    found = 0
    for sheet_name in PULL_SHEETS:
        ws = _find_worksheet(spreadsheet, sheet_name)
        is_mandatory = " (Bắt buộc)" if sheet_name in MANDATORY_SHEETS else ""
        if ws:
            _emit(logger, f"   ✅ {sheet_name}{is_mandatory}")
            found += 1
        else:
            icon = "❌" if sheet_name in MANDATORY_SHEETS else "⚠️"
            _emit(logger, f"   {icon} {sheet_name}{is_mandatory} — KHÔNG TÌM THẤY")
    _emit(logger, f"\n📈 Kết quả: {found}/{len(PULL_SHEETS)} sheet nguồn sẵn sàng")


def run_sync(
    config_path: str | Path = DEFAULT_CONFIG_PATH,
    logger: Logger | None = None,
) -> SyncResult:
    """Full pipeline: connect → pull → process → push → mirror."""
    start_time = time.time()
    result = SyncResult()

    _emit(logger, "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    _emit(logger, "☁️  HR-NEXUS Google Sheets Sync")
    _emit(logger, f"   Thời gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    _emit(logger, "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")

    try:
        # Step 1: Connect
        config = SyncConfig.load(config_path)
        spreadsheet = connect(config, logger)
        _emit(logger, "")

        # Step 2: Pull
        tables = pull_all_sources(spreadsheet, logger)
        result.sheets_pulled = len(tables)
        result.total_rows_pulled = sum(len(df) for df in tables.values())
        _emit(logger, "")

        # Step 3: Process
        outputs, source_updates = process(tables, config.fiscal_year, logger)
        _emit(logger, "")

        # Step 4: Push report outputs
        result.total_rows_pushed = push_all_outputs(spreadsheet, outputs, logger)
        result.sheets_pushed = len(outputs)
        _emit(logger, "")

        # Step 4b: Push source updates (raw_participants + training_records with QA status)
        if source_updates:
            _emit(logger, "─── PUSH SOURCE: Cập nhật QA status lên Google Sheets ───")
            for sheet_name, df in source_updates.items():
                try:
                    matrix = [df.columns.tolist()] + _sheets_safe_matrix(df.values.tolist())
                    pushed = _retry_with_backoff(
                        lambda sn=sheet_name, mx=matrix: push_sheet(spreadsheet, sn, mx, logger),
                        logger=logger,
                    )
                    result.total_rows_pushed += pushed
                    time.sleep(1.0)
                except Exception as exc:
                    _emit(logger, f"  ⚠️ Không thể cập nhật '{sheet_name}': {exc}")
            _emit(logger, "")

        # Step 5: Mirror to local Excel
        mirror_path = mirror_to_excel(
            outputs,
            source_updates=source_updates,
            backup=config.backup_before_push,
            logger=logger,
        )
        result.mirror_path = mirror_path
        _emit(logger, "")

    except SyncError as exc:
        result.success = False
        result.errors.append(f"[{exc.code}] {exc.friendly_message}")
        _emit(logger, f"\n❌ {exc.friendly_message}")
        if exc.__cause__:
            _emit(logger, f"   Chi tiết kỹ thuật: {exc.__cause__}")
    except Exception as exc:
        result.success = False
        result.errors.append(str(exc))
        _emit(logger, f"\n❌ {FRIENDLY_ERRORS['E04_NETWORK']}")
        _emit(logger, f"   Chi tiết: {exc}")

    result.duration_seconds = time.time() - start_time

    # Summary
    _emit(logger, "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    if result.success:
        _emit(
            logger,
            f"✅ Tuyệt vời! Báo cáo đã được cập nhật!\n"
            f"   📥 Kéo: {result.sheets_pulled} sheets ({result.total_rows_pulled} dòng)\n"
            f"   📤 Đẩy: {result.sheets_pushed} sheets ({result.total_rows_pushed} dòng)\n"
            f"   ⏱️  Thời gian: {result.duration_seconds:.1f} giây",
        )
    else:
        _emit(logger, f"❌ Đồng bộ thất bại. Lỗi: {'; '.join(result.errors)}")
    _emit(logger, "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

    return result


def run_pull_only(
    config_path: str | Path = DEFAULT_CONFIG_PATH,
    logger: Logger | None = None,
) -> None:
    """Pull data from Sheets → write to local Excel (no push back)."""
    config = SyncConfig.load(config_path)
    spreadsheet = connect(config, logger)
    tables = pull_all_sources(spreadsheet, logger)
    outputs, source_updates = process(tables, config.fiscal_year, logger)
    mirror_to_excel(outputs, source_updates=source_updates, backup=True, logger=logger)
    _emit(logger, "\n✅ Đã kéo data và cập nhật file Excel local.")


def run_push_only(
    config_path: str | Path = DEFAULT_CONFIG_PATH,
    logger: Logger | None = None,
) -> None:
    """Read local Excel → push results to Google Sheets."""
    config = SyncConfig.load(config_path)
    spreadsheet = connect(config, logger)
    _emit(logger, "Đang đọc dữ liệu từ file Excel local...")
    inputs = fetch_data(DEFAULT_WORKBOOK_PATH, logger=lambda msg: _emit(logger, f"  {msg}"))
    outputs = transform_data(inputs, fiscal_year=config.fiscal_year)
    push_all_outputs(spreadsheet, outputs, logger)
    _emit(logger, "\n✅ Đã đẩy kết quả từ Excel local lên Google Sheets.")


def run_push_full(
    workbook_path: str | Path = DEFAULT_WORKBOOK_PATH,
    config_path: str | Path = DEFAULT_CONFIG_PATH,
    logger: Logger | None = None,
) -> SyncResult:
    """Luồng 3: Đọc toàn bộ sheets từ Excel local → ghi đè lên Google Sheets.

    Không tính toán, không transform — chỉ đọc nguyên trạng và đẩy lên.
    Dùng openpyxl (không pandas) để tối ưu bộ nhớ.
    """
    start_time = time.time()
    result = SyncResult()

    _emit(logger, "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    _emit(logger, "📤  HR-NEXUS: Đẩy Excel Local → Google Sheets")
    _emit(logger, f"   Thời gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    _emit(logger, "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")

    try:
        # ── Step 1: Mở file Excel local (E02: PermissionError / FileNotFound) ──
        wb_path = Path(workbook_path).expanduser()
        if not wb_path.is_absolute():
            wb_path = (Path.cwd() / wb_path).resolve()

        _emit(logger, f"─── Đang đọc file Excel: {wb_path.name} ───")

        try:
            workbook = _openpyxl_load_workbook(
                str(wb_path), read_only=True, data_only=True,
            )
        except PermissionError:
            raise SyncError(
                "E02_FILE_LOCKED",
                "Không thể đọc file Excel vì nó đang được mở. "
                "Vui lòng đóng file Excel lại và thử nhé.",
            )
        except FileNotFoundError:
            raise SyncError(
                "E02_FILE_LOCKED",
                f"Không tìm thấy file Excel: {wb_path}",
            )

        try:
            sheet_names = workbook.sheetnames
            _emit(logger, f"  📋 Tìm thấy {len(sheet_names)} sheets trong file Excel\n")

            # ── Step 2: Kết nối Google Sheets ──
            config = SyncConfig.load(config_path)
            spreadsheet = connect(config, logger)
            _emit(logger, "")

            # ── Step 3: Đọc + đẩy từng sheet (E03: retry on API error) ──
            _emit(logger, "─── PUSH FULL: Đang ghi toàn bộ lên Google Sheets ───")
            total_pushed = 0
            sheets_ok = 0

            for idx, sheet_name in enumerate(sheet_names, 1):
                _emit(logger, f"  [{idx}/{len(sheet_names)}] {sheet_name}...")

                worksheet = workbook[sheet_name]
                matrix = _read_excel_sheet_as_matrix(worksheet)

                if not matrix:
                    _emit(logger, "    ⏭️ Sheet rỗng, bỏ qua.")
                    continue

                safe_matrix = _sheets_safe_matrix(matrix)

                try:
                    pushed = _retry_with_backoff(
                        lambda sn=sheet_name, mx=safe_matrix: push_sheet(
                            spreadsheet, sn, mx, logger,
                        ),
                        logger=logger,
                    )
                    total_pushed += pushed or 0
                    sheets_ok += 1
                except SyncError as exc:
                    msg = (
                        f"Lỗi kết nối khi đang đẩy dữ liệu lên '{sheet_name}'. "
                        "Vui lòng kiểm tra mạng và thử lại."
                    )
                    result.warnings.append(msg)
                    _emit(logger, f"    ⚠️ {msg}")

                # Rate-limit cooldown giữa các sheets
                if idx < len(sheet_names):
                    time.sleep(1.0)

            result.sheets_pushed = sheets_ok
            result.total_rows_pushed = total_pushed

        finally:
            workbook.close()

    except SyncError as exc:
        result.success = False
        result.errors.append(f"[{exc.code}] {exc.friendly_message}")
        _emit(logger, f"\n❌ {exc.friendly_message}")
        if exc.__cause__:
            _emit(logger, f"   Chi tiết kỹ thuật: {exc.__cause__}")
    except Exception as exc:
        result.success = False
        result.errors.append(str(exc))
        _emit(logger, f"\n❌ Lỗi không mong đợi: {exc}")

    result.duration_seconds = time.time() - start_time

    # ── Summary ──
    _emit(logger, "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    if result.success:
        _emit(
            logger,
            f"✅ Tuyệt vời! Đã đẩy toàn bộ Excel lên Google Sheets!\n"
            f"   📤 Đã ghi: {result.sheets_pushed} sheets ({result.total_rows_pushed} dòng)\n"
            f"   ⏱️  Thời gian: {result.duration_seconds:.1f} giây",
        )
    else:
        _emit(logger, f"❌ Đẩy dữ liệu thất bại. Lỗi: {'; '.join(result.errors)}")
    _emit(logger, "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

    return result


# ── Helpers ──────────────────────────────────────────────────────────────────


def _find_worksheet(
    spreadsheet: gspread.Spreadsheet,
    expected_name: str,
) -> gspread.Worksheet | None:
    """Find worksheet by normalized name (tolerant matching)."""
    target = _normalize_token_simple(expected_name)
    for ws in spreadsheet.worksheets():
        if _normalize_token_simple(ws.title) == target:
            return ws
    return None


def _retry_with_backoff(
    func: Callable,
    max_retries: int = 3,
    base_delay: float = 2.0,
    logger: Logger | None = None,
) -> Any:
    """Retry with exponential backoff for API rate limits (E07)."""
    for attempt in range(1, max_retries + 1):
        try:
            return func()
        except gspread.exceptions.APIError as exc:
            if exc.response.status_code == 429 and attempt < max_retries:
                delay = base_delay * (2 ** (attempt - 1))
                _emit(
                    logger,
                    f"  ⏳ {FRIENDLY_ERRORS['E07_RATE_LIMIT'].format(attempt=attempt)}",
                )
                time.sleep(delay)
            else:
                raise
        except SyncError:
            raise
        except (ConnectionError, TimeoutError, OSError) as exc:
            if attempt < max_retries:
                delay = base_delay * (2 ** (attempt - 1))
                _emit(logger, f"  ⏳ Mất kết nối, thử lại sau {delay:.0f}s... (lần {attempt}/3)")
                time.sleep(delay)
            else:
                raise SyncError("E04_NETWORK", FRIENDLY_ERRORS["E04_NETWORK"], cause=exc)
    return None


def _sheets_safe_matrix(matrix: list[list[Any]]) -> list[list[Any]]:
    """Convert matrix values to types safe for Google Sheets API."""
    safe = []
    for row in matrix:
        safe_row = []
        for val in row:
            if val is None:
                safe_row.append("")
            elif pd.isna(val):
                safe_row.append("")
            elif isinstance(val, pd.Timestamp):
                safe_row.append(val.strftime("%Y-%m-%d"))
            elif isinstance(val, datetime):
                safe_row.append(val.strftime("%Y-%m-%d"))
            elif hasattr(val, "item") and callable(val.item):
                try:
                    safe_row.append(val.item())
                except Exception:
                    safe_row.append(str(val))
            else:
                safe_row.append(val)
        safe.append(safe_row)
    return safe


def _emit(logger: Logger | None, message: str) -> None:
    if logger is not None:
        logger(message)


def _read_excel_sheet_as_matrix(worksheet) -> list[list[Any]]:
    """Read an openpyxl worksheet into a trimmed list-of-lists matrix.

    Uses openpyxl only (no pandas) for memory efficiency.
    Trims trailing blank rows and columns.
    """
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    # Determine bounding box: last non-blank row and max column
    last_data_row = -1
    max_col = 0
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell is not None and str(cell).strip() != "":
                last_data_row = i
                if j + 1 > max_col:
                    max_col = j + 1

    if last_data_row < 0:
        return []

    # Build trimmed matrix
    matrix: list[list[Any]] = []
    for row in rows[: last_data_row + 1]:
        trimmed = list(row[:max_col])
        # Pad short rows to uniform width
        while len(trimmed) < max_col:
            trimmed.append(None)
        matrix.append(trimmed)

    return matrix


# ── CLI ──────────────────────────────────────────────────────────────────────


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="HR-NEXUS Google Sheets ↔ Local Python Sync Bridge.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Ví dụ:\n"
            "  python gsheet_sync.py check       Kiểm tra kết nối\n"
            "  python gsheet_sync.py sync        Đồng bộ đầy đủ\n"
            "  python gsheet_sync.py pull-only   Chỉ kéo data về Excel\n"
            "  python gsheet_sync.py push-only   Chỉ đẩy từ Excel lên Sheets\n"
            "  python gsheet_sync.py push-full   Đẩy toàn bộ Excel ghi đè Sheets\n"
        ),
    )
    parser.add_argument(
        "action",
        choices=("check", "sync", "pull-only", "push-only", "push-full"),
        nargs="?",
        default="sync",
        help="Hành động cần thực hiện (mặc định: sync).",
    )
    parser.add_argument(
        "--config",
        default=str(DEFAULT_CONFIG_PATH),
        help="Đường dẫn file sync_config.json.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        pass

    parser = _build_parser()
    args = parser.parse_args(argv)

    actions = {
        "check": run_check,
        "sync": run_sync,
        "pull-only": run_pull_only,
        "push-only": run_push_only,
        "push-full": run_push_full,
    }

    try:
        action_fn = actions[args.action]
        result = action_fn(config_path=args.config, logger=print)
        if args.action == "sync" and isinstance(result, SyncResult) and not result.success:
            return 1
        return 0
    except SyncError as exc:
        print(f"\n❌ {exc.friendly_message}")
        return 1
    except Exception as exc:
        print(f"\n❌ {FRIENDLY_ERRORS['E04_NETWORK']}")
        print(f"   Chi tiết: {exc}")
        if "--debug" in (argv or []):
            traceback.print_exc()
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
