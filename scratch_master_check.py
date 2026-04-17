import pandas as pd
from openpyxl import load_workbook
import warnings
warnings.simplefilter(action='ignore', category=UserWarning)

path = 'o:/clean-architecture-canonical/LD REPORT V2/LD REPORT V2/file excel tren gg sheet/L&D - 2026 MASTER DATA - TEST v4.xlsx'
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb['Danh mục nhân viên']
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
df = pd.DataFrame(rows[1:], columns=headers)

df = df.dropna(how='all')
if 'Trạng thái' in df.columns:
    df = df[~df['Trạng thái'].str.lower().eq('inactive')]

id_col = next((c for c in df.columns if 'mã' in str(c).lower() and 'nhân viên' in str(c).lower()), df.columns[0])
name_col = next((c for c in df.columns if 'họ' in str(c).lower() and 'tên' in str(c).lower() or 'tên' in str(c).lower() and 'nhân viên' in str(c).lower()), df.columns[1])
email_col = next((c for c in df.columns if 'email' in str(c).lower()), df.columns[2])

def clean(val):
    if pd.isna(val): return ""
    if isinstance(val, float) and val.is_integer(): return str(int(val))
    return str(val).strip()

df[id_col] = df[id_col].apply(clean)
df[email_col] = df[email_col].apply(clean).str.lower()
df[name_col] = df[name_col].apply(clean)

with open('scratch_output.txt', 'w', encoding='utf-8') as f:
    f.write("--- 1. KIEM TRA TRUNG MA NHAN VIEN ---\n")
    dup_ids = df[df[id_col].ne("") & df.duplicated(subset=[id_col], keep=False)]
    if not dup_ids.empty:
        for _, row in dup_ids.sort_values(id_col).iterrows():
            f.write(f"Ma NV: {row[id_col]} | Ten: {row[name_col]} | Email: {row[email_col]}\n")
    else:
        f.write("SACH! Khong co ma nhan vien bi trung.\n")

    f.write("\n--- 2. KIEM TRA TRUNG EMAIL ---\n")
    dup_emails = df[df[email_col].ne("") & df.duplicated(subset=[email_col], keep=False)]
    if not dup_emails.empty:
        for _, row in dup_emails.sort_values(email_col).iterrows():
            f.write(f"Email: {row[email_col]} | Ma NV: {row[id_col]} | Ten: {row[name_col]}\n")
    else:
        f.write("SACH! Khong co Email bi trung.\n")

    f.write("\n--- 3. CAC DONG CO CHUONG.LE ---\n")
    chuong_rows = df[df[email_col].str.contains('chuong.le', na=False)]
    for _, row in chuong_rows.iterrows():
        f.write(f"Email: {row[email_col]} | Ma NV: {row[id_col]} | Ten: {row[name_col]}\n")
